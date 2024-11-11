"""Модуль для форматирования данных о недвижимости, сохранения их в Excel и вывода в консоль в виде таблицы.
"""

import logging
import os
import textwrap
from datetime import datetime, timezone
from typing import Optional, Dict, Any, List, Tuple

import openpyxl
import pandas as pd
from prettytable import PrettyTable, ALL, DOUBLE_BORDER

logger = logging.getLogger(__name__)


def format_rights(right: Dict[str, Any]) -> str:
    """
    Форматирует информацию о правах на недвижимость.

    :param right: Словарь с информацией о праве на недвижимость.
    :type right: Dict[str, Any]
    :return: Строка с отформатированной информацией о праве.
    :rtype: str
    """
    # Логируем начало форматирования
    logger.debug("Formatting rights data.")

    # Извлекаем тип права
    right_type: str = right.get('rightTypeDesc', '')
    # Извлекаем номер права
    right_number: str = right.get('rightNumber', '')
    # Извлекаем и форматируем дату регистрации права
    right_reg_date: Optional[str] = get_normal_date(right.get('rightRegDate'))

    # Формируем отформатированную строку
    formatted_right: str = f"{right_type}, {right_number} от {right_reg_date}"

    return formatted_right


def format_encumbrances(encumbrance: Dict[str, Any]) -> str:
    """
    Форматирует информацию об обременениях недвижимости.

    :param encumbrance: Словарь с информацией об обременении.
    :type encumbrance: Dict[str, Any]
    :return: Строка с отформатированной информацией об обременении.
    :rtype: str
    """
    # Логируем начало форматирования
    logger.debug("Formatting encumbrance data.")

    # Извлекаем тип обременения
    type_desc: str = encumbrance.get('typeDesc', '')
    # Извлекаем номер обременения
    encumbrance_number: str = encumbrance.get('encumbranceNumber', '')
    # Извлекаем дату начала обременения, если она существует
    start_date_millis: Optional[int] = encumbrance.get('startDate')
    start_date: Optional[str] = get_normal_date(start_date_millis) if start_date_millis else ''

    # Формируем отформатированную строку
    formatted_encumbrance: str = f"{type_desc} {encumbrance_number}"
    if start_date:
        formatted_encumbrance += f" от {start_date}"

    return formatted_encumbrance


def get_normal_date(millis: Optional[int]) -> Optional[str]:
    """
    Преобразует миллисекунды в строку даты формата 'дд.мм.гггг'.

    :param millis: Время в миллисекундах.
    :type millis: Optional[int]
    :return: Форматированная дата в виде строки или None, если входное значение None.
    :rtype: Optional[str]
    """
    # Логируем начало преобразования даты
    logger.debug("Converting milliseconds to date string.")

    # Проверяем, является ли входное значение None
    if millis is None:
        return None

    # Преобразуем миллисекунды в секунды
    seconds: float = millis / 1000.0

    # Получаем объект даты и времени
    date_object: datetime = datetime.fromtimestamp(seconds, tz=timezone.utc)

    # Форматируем дату в строку 'дд.мм.гггг'
    formatted_date: str = date_object.strftime('%d.%m.%Y')

    return formatted_date


def save_to_excel(result_df: pd.DataFrame, filename: str) -> None:
    """
    Сохраняет данные из DataFrame в файл Excel с применением форматирования.

    :param result_df: DataFrame с данными для сохранения.
    :type result_df: pd.DataFrame
    :param filename: Имя выходного файла без расширения.
    :type filename: str
    :return: None
    :rtype: None
    """
    # Логируем начало процесса сохранения
    logger.info("Saving data to Excel...")

    # Создаем директорию 'output', если она не существует
    os.makedirs('output', exist_ok=True)

    # Открываем ExcelWriter и сохраняем данные в файл Excel
    with pd.ExcelWriter(f'output/{filename}.xlsx', engine='openpyxl') as writer:
        # Записываем DataFrame в Excel без индексов строк
        result_df.to_excel(writer, index=False)

        # Получаем лист 'Sheet1' для дальнейшего форматирования
        worksheet: openpyxl.worksheet.worksheet.Worksheet = writer.sheets['Sheet1']

        # Применяем форматирование к листу
        format_worksheet(worksheet, result_df)


def format_worksheet(worksheet: openpyxl.worksheet.worksheet.Worksheet, result_df: pd.DataFrame) -> None:
    """
    Форматирует лист Excel, выравнивая текст и устанавливая ширину колонок для заданных столбцов.

    :param worksheet: Лист Excel, который нужно отформатировать.
    :type worksheet: openpyxl.worksheet.worksheet.Worksheet
    :param result_df: DataFrame с данными для определения колонок и их ширины.
    :type result_df: pd.DataFrame
    :return: None
    :rtype: None
    """
    # Логируем начало процесса форматирования листа
    logger.info("Formatting Excel worksheet: aligning text and setting column widths.")

    # Список колонок, для которых нужно применить перенос текста
    wrap_text_columns: List[str] = [
        'Адрес',
        'Категория земель',
        'Вид, номер и дата государственной регистрации права',
        'Ограничение прав и обременение объекта недвижимости',
        'Тип объекта',
        'Вид разрешенного использования'
    ]

    # Итерация по строкам листа, начиная со 2-й строки
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                   min_col=1, max_col=worksheet.max_column):
        format_row(row, result_df, wrap_text_columns)

    # Вызываем функцию для установки ширины колонок
    set_column_widths(worksheet, result_df)


def format_row(row: Tuple[openpyxl.cell.cell.Cell, ...], result_df: pd.DataFrame,
               wrap_text_columns: List[str]) -> None:
    """
    Форматирует строку листа Excel, устанавливая выравнивание и перенос текста для каждой ячейки.

    :param row: Строка листа Excel.
    :type row: Tuple[openpyxl.cell.cell.Cell, ...]
    :param result_df: DataFrame с данными для определения колонок.
    :type result_df: pd.DataFrame
    :param wrap_text_columns: Список названий колонок, для которых требуется перенос текста.
    :type wrap_text_columns: List[str]
    :return: None
    :rtype: None
    """
    for cell in row:
        # Определяем индекс и название колонки
        column_index: int = cell.column - 1

        # Проверяем корректность индекса колонки
        if column_index >= len(result_df.columns):
            continue  # Ранний выход, если индекс вне диапазона

        column_name: str = result_df.columns[column_index]

        # Ранний выход, если название колонки отсутствует
        if not column_name:
            continue

        # Устанавливаем выравнивание текста по умолчанию
        alignment: openpyxl.styles.Alignment = openpyxl.styles.Alignment(
            vertical='top', horizontal='left'
        )

        # Ранний переход к следующей ячейке, если колонка не требует переноса текста
        if column_name not in wrap_text_columns:
            cell.alignment = alignment
            continue

        # Устанавливаем wrap_text=True для колонок, требующих переноса текста
        alignment = openpyxl.styles.Alignment(
            vertical='top', horizontal='left', wrap_text=True
        )

        # Применяем выравнивание к ячейке
        cell.alignment = alignment


def set_column_widths(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                      result_df: pd.DataFrame) -> None:
    """
    Устанавливает ширину колонок в листе Excel на основе заданных значений.

    :param worksheet: Лист Excel, для которого нужно установить ширину колонок.
    :type worksheet: openpyxl.worksheet.worksheet.Worksheet
    :param result_df: DataFrame с данными для определения названий колонок.
    :type result_df: pd.DataFrame
    :return: None
    :rtype: None
    """
    # Словарь с названиями колонок и их шириной
    column_widths: Dict[str, int] = {
        'Кадастровый номер': 19,
        'Адрес': 45,
        'Категория земель': 22,
        'Вид, номер и дата государственной регистрации права': 45,
        'Ограничение прав и обременение объекта недвижимости': 45
    }

    for col_name, width in column_widths.items():
        # Ранний переход, если колонка отсутствует в DataFrame
        if col_name not in result_df.columns:
            continue

        # Получаем индекс колонки по её имени
        col_index: int = result_df.columns.get_loc(col_name) + 1
        col_letter: str = openpyxl.utils.get_column_letter(col_index)

        # Устанавливаем ширину колонки
        worksheet.column_dimensions[col_letter].width = width


def print_pretty_table(df: pd.DataFrame, wrap_width: int = 50) -> None:
    """
    Принимает DataFrame, переворачивает его, создаёт PrettyTable в стиле DOUBLE_BORDER и выводит на экран.

    :param df: DataFrame для форматирования.
    :type df: pd.DataFrame
    :param wrap_width: Ширина строки для переноса длинного текста (по умолчанию 50 символов).
    :type wrap_width: int
    :return: None
    :rtype: None
    """
    # Логируем начало процесса вывода таблицы
    logger.info("Printing PrettyTable from DataFrame.")

    # Переворачиваем DataFrame и сбрасываем индексы
    df_transposed: pd.DataFrame = df.T.reset_index()

    # Обновляем названия колонок
    df_transposed.columns = ['Column_Headers'] + [f'Row_{i}' for i in range(1, len(df_transposed.columns))]

    # Создаём PrettyTable с нужными настройками
    table: PrettyTable = PrettyTable()
    table.header = False  # Отключаем заголовок таблицы
    table.align = "l"  # Выравниваем текст по левому краю
    table.hrules = ALL  # Добавляем линии между строками
    table.set_style(DOUBLE_BORDER)  # Устанавливаем стиль DOUBLE_BORDER

    # Добавляем строки в таблицу с переносом длинного текста
    for row in df_transposed.itertuples(index=False):
        # Перенос строк на указанный wrap_width
        wrapped_row: List[str] = [textwrap.fill(str(cell), width=wrap_width) for cell in row]
        table.add_row(wrapped_row)

    # Выводим таблицу в консоль
    print(table)
