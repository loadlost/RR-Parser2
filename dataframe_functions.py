# dataframe_functions.py
# Модуль для работы с данными: добавление, очистка и форматирование для сохранения в Excel.
# Включает функции для обработки DataFrame, управления стилями и сохранения результатов.
# Ориентирован на подготовку и представление данных, полученных из различных источников.

import logging

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

import pandas as pd

from config import columns
from requests_functions import add_to_dataframe

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
result_df = pd.DataFrame(columns=columns)  # Создание пустого DataFrame с колонками, определенными в 'columns'.


def add_data_to_dataframe(response):
    # Добавление данных из ответа в DataFrame.
    # Извлекает элементы из JSON-ответа и добавляет их в DataFrame 'result_df'.
    # Аргументы:
    #   response (Response): Объект Response библиотеки requests, содержащий JSON-ответ.
    # Возвращает:
    #   None

    logger.info("Добавление данных в DataFrame...")

    # Извлечение элементов из JSON-ответа
    result = response.json()
    elements = result.get('elements', [])

    if elements:
        # Если есть элементы, добавляем каждый из них в DataFrame
        for element in elements:
            data = add_to_dataframe(element)
            global result_df
            result_df = pd.concat([result_df, pd.DataFrame([data])], ignore_index=True)
    else:
        logger.warning("Объекты не найдены")

    logger.info("Добавление данных завершено.")


def clean_dataframe():
    # Удаление пустых колонок в DataFrame 'result_df'.
    # Удаляет колонки, в которых все значения NaN или пустые.
    # Входные данные: отсутствуют.
    # Выходные данные: отсутствуют, но DataFrame очищается от пустых колонок.

    logger.info("Удаление пустых колонок в DataFrame...")

    # Удаление колонок, где все значения являются NaN (пустые)
    result_df.dropna(axis=1, how='all', inplace=True)


def format_worksheet(worksheet):
    # Функция форматирования листа Excel.
    # Выравнивает текст и устанавливает ширину колонок для переданного листа.
    # Аргументы:
    #   worksheet: Лист Excel, который нужно отформатировать.
    # Входные данные:
    #   worksheet (Worksheet): Объект листа Excel.
    # Выходные данные: отсутствуют, но лист форматируется.

    logger.info("Выравнивание текста и установка ширины колонок в Excel...")

    # Итерация по строкам листа, начиная с 2-й строки (первая строка - заголовок)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            # Установка выравнивания текста по вертикали и горизонтали (по умолчанию - сверху и слева)
            alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left')

            # Устанавливаем wrap_text=True для определенных колонок, чтобы текст переносился на новую строку
            if result_df.columns[cell.column - 1] in ['Адрес', 'Категория земель',
                                                      'Вид, номер и дата государственной регистрации права',
                                                      'Ограничение прав и обременение объекта недвижимости']:
                alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left', wrap_text=True)

            cell.alignment = alignment

    # Задание ширины колонок
    column_widths = {
        'Кадастровый номер': 19,
        'Адрес': 45,
        'Категория земель': 22,
        'Вид, номер и дата государственной регистрации права': 45,
        'Ограничение прав и обременение объекта недвижимости': 45
    }

    for col_name, width in column_widths.items():
        # Получение номера колонки по ее имени
        col_letter = pd.Series(result_df.columns).eq(col_name).idxmax() + 1

        # Установка ширины колонки
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_letter)].width = width


def save_to_excel():
    # Функция сохранения данных в Excel файл.
    # Создает или обновляет файл 'output_data.xlsx' с данными DataFrame.
    # Входные данные: отсутствуют.
    # Выходные данные: отсутствуют, но данные сохраняются в файл Excel.

    logger.info("Сохранение данных в Excel...")

    # Используем pd.ExcelWriter для сохранения в существующий файл
    with pd.ExcelWriter('output_data.xlsx', engine='openpyxl') as writer:
        # Записываем данные в Excel, без индексов строк
        result_df.to_excel(writer, index=False)

        # Получаем лист 'Sheet1'
        worksheet = writer.sheets['Sheet1']

        # Применяем форматирование к листу
        format_worksheet(worksheet)
